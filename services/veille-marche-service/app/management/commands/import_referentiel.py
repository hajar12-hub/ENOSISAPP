"""Importe le référentiel produit depuis le classeur livré dans ``data/``.

Le format est celui du Price Ladder : les marques sont sur la deuxième ligne
et leurs SKU sur la troisième. La commande est idempotente et peut donc être
rejouée après chaque déploiement sans dupliquer les références.
"""

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from app.models import Categorie, Concurent, Marque, Segment, SKU, SMarque, SousCategorie


DATA_FILE = Path(__file__).resolve().parents[5] / "data" / "Price ladder PAPER Tracker (1) (1).xlsx"


class Command(BaseCommand):
    help = "Importe les catégories, sous-catégories, segments, marques et SKU depuis data/."

    def handle(self, *args, **options):
        if not DATA_FILE.exists():
            raise CommandError(f"Classeur introuvable : {DATA_FILE}")

        workbook = load_workbook(DATA_FILE, read_only=True, data_only=True)
        detergent = Categorie.objects.filter(nom="Detergent").order_by("created_at").first()
        if not detergent:
            detergent = Categorie.objects.create(nom="Detergent")
        paper = Categorie.objects.filter(nom="Papier").order_by("created_at").first()
        if not paper:
            paper = Categorie.objects.create(nom="Papier")

        self._ensure_detergent_subcategories(detergent)
        self._import_high_suds(workbook["Sheet1"], detergent)
        self._import_other_detergents(detergent)
        for sheet_name, sous_categorie_name in {
            "TOILET PAPER ": "Toilet Paper",
            "FACIAL TISSUES": "Facial Tissues",
            "PAPER TOWEL": "Paper Towel",
        }.items():
            self._import_paper_sheet(workbook[sheet_name], paper, sous_categorie_name)

        self.stdout.write(self.style.SUCCESS("Référentiel importé depuis data/ avec succès."))

    @staticmethod
    def _text(value):
        return str(value).strip() if value is not None and str(value).strip() else None

    def _import_high_suds(self, sheet, categorie):
        sous_categorie = self._subcategory("High Suds", categorie, SousCategorie.Type.DIRECT)
        current_brand = None
        for column in range(1, sheet.max_column + 1):
            brand_name = self._text(sheet.cell(2, column).value)
            if brand_name:
                current_brand = brand_name
            sku_name = self._text(sheet.cell(3, column).value)
            if not current_brand or not sku_name:
                continue
            marque, _ = Marque.objects.get_or_create(
                nom=current_brand,
                defaults={"concurrent": current_brand.upper() != "MIO"},
            )
            SMarque.objects.get_or_create(marque=marque, sous_categorie=sous_categorie)
            SKU.objects.get_or_create(nom=sku_name, marque=marque)

    @staticmethod
    def _subcategory(nom, categorie, type_):
        sous_categorie = SousCategorie.objects.filter(nom=nom, categorie=categorie).order_by("created_at").first()
        return sous_categorie or SousCategorie.objects.create(nom=nom, categorie=categorie, type=type_)

    @classmethod
    def _ensure_detergent_subcategories(cls, categorie):
        for nom, type_ in (
            ("High Suds", SousCategorie.Type.DIRECT),
            ("Low Suds Liquide", SousCategorie.Type.GROUPS_INLINE),
            ("Liquide Vaisselle", SousCategorie.Type.DIRECT),
            ("Poudre Matic", SousCategorie.Type.DIRECT),
        ):
            cls._subcategory(nom, categorie, type_)

    @staticmethod
    def _import_other_detergents(categorie):
        references = {
            "Liquide Vaisselle": {
                "MIO": ["300ml", "750ml", "1250ml"],
                "ONI": ["300ml", "750ml", "1250ml"],
                "MAXIS": ["300ml", "750ml", "1250ml"],
                "ARIEL": ["900ml"],
            },
            "Poudre Matic": {
                "MIO": ["400 GR", "750+150GR"],
                "ARIEL": ["400gr", "750gr"],
                "MAXIS": ["400 gr", "900 gr"],
            },
        }
        for sous_categorie_nom, brands in references.items():
            sous_categorie = SousCategorie.objects.filter(
                nom=sous_categorie_nom, categorie=categorie
            ).order_by("created_at").first()
            for brand_name, sku_names in brands.items():
                marque, _ = Marque.objects.get_or_create(
                    nom=brand_name, defaults={"concurrent": brand_name != "MIO"}
                )
                SMarque.objects.get_or_create(marque=marque, sous_categorie=sous_categorie)
                for sku_name in sku_names:
                    SKU.objects.get_or_create(nom=sku_name, marque=marque)

        low_suds = SousCategorie.objects.filter(nom="Low Suds Liquide", categorie=categorie).order_by("created_at").first()
        for brand_name in ("MIO", "ARIEL"):
            marque, _ = Marque.objects.get_or_create(
                nom=brand_name, defaults={"concurrent": brand_name != "MIO"}
            )
            Concurent.objects.get_or_create(
                nom="vs MIO", marque=marque, sous_categorie=low_suds
            )
            for sku_name in ("1L", "1,8L", "3L"):
                SKU.objects.get_or_create(nom=sku_name, marque=marque)

    def _import_paper_sheet(self, sheet, categorie, sous_categorie_name):
        sous_categorie = self._subcategory(sous_categorie_name, categorie, SousCategorie.Type.SEGMENT)
        current_segment = None
        current_brand = None
        for column in range(1, sheet.max_column + 1):
            segment_name = self._text(sheet.cell(1, column).value)
            if segment_name:
                current_segment = segment_name
            brand_name = self._text(sheet.cell(2, column).value)
            if brand_name and brand_name.lower() != "paper":
                current_brand = brand_name
            sku_name = self._text(sheet.cell(3, column).value)
            if not current_segment or not current_brand or not sku_name:
                continue
            segment, _ = Segment.objects.get_or_create(
                nom=current_segment, sous_categorie=sous_categorie
            )
            marque, _ = Marque.objects.get_or_create(
                nom=current_brand,
                segment=segment,
                defaults={"concurrent": current_brand.upper() != "PAPILLON"},
            )
            SKU.objects.get_or_create(nom=sku_name, marque=marque)
